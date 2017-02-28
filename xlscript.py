import sys
from openpyxl import load_workbook
from openpyxl import Workbook

weekday_col_list = [
'L', 'N', 'P', 'R', 'T', 'V', 'X', 'Z', 'AD', 'AF','AH', 'AJ','AL','AN', 'AP', 'AR', 'AW', 
'AY', 'BA', 'BC', 'BE', 'BG', 'BI', 'BK', 'BN', 'BP', 'BR', 'BT', 'BV', 'BX','BZ', 'CB', 'CD', 
'CF', 'CH', 'CJ', 'CL', 'CN', 'CP', 'CR', 'CU', 
'CW', 'CY','DA', 'DC', 'DE', 'DG', 'DI', 'DM', 
'DO', 'DQ', 'DS', 'DU', 'DW', 'DY', 'EA','EF', 
'EH', 'EJ', 'EL', 'EN', 'EP', 'ER', 'ET']

weekend_col_list = [
'K', 'M', 'O', 'Q', 'S', 'U', 'W', 'Y', 'AD', 'AF', 'AH', 'AJ', 'AL', 'AN',
'AP', 'AR', 'AU', 'AW', 'AY', 'BA', 'BC', 'BE', 'BG', 'BI', 'BK', 'BM', 'BO',
'BQ', 'BS', 'BU', 'BW', 'BY', 'CB', 'CD', 'CF', 'CH', 'CJ', 'CL', 'CN', 'CP',
'CT', 'CV', 'CX', 'CZ', 'DB', 'DD', 'DF', 'DH', 'DM', 'DO', 'DQ', 'DS', 'DU',
'DW', 'DY', 'EA']

avg_di

'5-15 min': 10,
'15-30 min': 22.5,
'30-45 min': 37.5,
'45-60 min': 52.5,
'1-1.5 h': 75,
'1.5-2 h': 105,
'2+': 120,
'2-2.5 h': 135,
'2.5-3 h': 165,
'3+ h': 180
}

def subCol(col_str):
	if len(col_str) == 1 and col_str != 'A':
		return chr(ord(col_str) - 1)
	else:
		if col_str[-1] == 'A':
			first = chr(ord(col_str[0]) - 1)
			return first + 'Z'
		else:
			last = chr(ord(col_str[-1]) - 1)
			return col_str[0] + last

def getColNum(col_str):
	if len(col_str) == 1:
		return (ord(col_str) % ord('A')) + 1
	else: 
		first = (ord(col_str[0]) % ord('A')) + 1
		second = (ord(col_str[-1]) % ord('A')) + 1
		col_num = 26 * first
		col_num += second
		return col_num

#command line params: 1-filename, 2-num_rows, 3 "week" or "weekend"
def main(argv):
	filename = argv[1]
	num_rows = int(argv[2])
	wb = load_workbook(filename)
	ws = wb.active
    
    #parse the final command line element to determine which column list to use
    if argv[3] == 'week':
        col_list = weekday_col_list
    elif argv[3] == 'weekend':
        col_list = weekend_col_list
    else:
        print 'You must specify week or weekend as final argument'
        return

	for col in col_list:
		start_str = col + '1'
		end_str = col + '{0}'.format(num_rows)
		cell_range = ws[start_str:end_str]
		for cell in cell_range:
			prev = ws.cell(row = cell[0].row, column = getColNum(subCol(cell[0].column)))
			if prev.value in avg_dict:
				c = ws.cell(row = cell[0].row, column = getColNum(cell[0].column))
				c.value = avg_dict[prev.value]
	wb.save(filename)
			
if __name__ == '__main__':
	main(sys.argv)
